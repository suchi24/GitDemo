import openpyxl

book = openpyxl.load_workbook("H:\\Python\\pythonExcelData.xlsx")

sheet = book.active
Dict = {}

cell = sheet.cell(row=1,column=2)

print(cell.value)

sheet.cell(row=2, column=2).value = "Suchi"

print(sheet.cell(row=2, column=2).value)

print("max rows ", sheet.max_row)
print("max cols ", sheet.max_column)

#print(sheet['A3'].value)
max_row = sheet.max_row

for i in range(1,max_row+1 ):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2,sheet.max_column+1):
            #print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)