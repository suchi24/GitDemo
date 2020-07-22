import openpyxl
class HomePageData:
    test_HomePage_data = [{"firstname" : "Suchi","email" : "abc@gmail.com", "gender" : "Female"},
                            {"firstname" : "sarva","email" : "134@gmail.com", "gender" : "Male"}]
    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("H:\\Python\\pyExcelData.xlsx")

        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]